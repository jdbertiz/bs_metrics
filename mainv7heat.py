import os
import csv
from openpyxl import load_workbook
from collections import defaultdict
import matplotlib.pyplot as plt
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
from datetime import datetime
import seaborn as sns

# Path to the BSREPORTS folder
folder_path = 'BSREPORTS'

# Dictionary to store aggregated results
aggregated_data = defaultdict(lambda: [0, 0])  # key: (Content, Type), value: [Unique Viewers sum, Viewers sum]

# Loop through all files in the folder
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(folder_path, filename)
        print(f"\nProcessing file: {file_path}")
        
        workbook = load_workbook(filename=file_path)
        sheet_names = workbook.sheetnames
        print("Sheets in the workbook:", sheet_names)
        
        for sheet_name in sheet_names:
            if sheet_name.lower() == "popular content":
                sheet = workbook[sheet_name]
                print(f"\nSheet: {sheet_name}")
                
                # Start reading from row 7
                for row in sheet.iter_rows(min_row=7, values_only=True):
                    if all(cell is None for cell in row):
                        break
                    
                    cleaned_row = [cell for cell in row if cell is not None]
                    
                    if len(cleaned_row) >= 4:
                        content = cleaned_row[0]
                        page_type = cleaned_row[1]
                        try:
                            unique_viewers = int(cleaned_row[2])
                            viewers = int(cleaned_row[3])
                        except ValueError:
                            continue  # Skip if values are not numbers
                        
                        key = (content, page_type)
                        aggregated_data[key][0] += unique_viewers
                        aggregated_data[key][1] += viewers

# Prepare data for display and chart
data = []
type_totals = defaultdict(int)

for (content, page_type), (unique_viewers, viewers) in aggregated_data.items():
    data.append([content, page_type, unique_viewers, viewers])
    type_totals[page_type] += viewers

# Convert to DataFrame for clean table output
df = pd.DataFrame(data, columns=['Content', 'Type', 'Unique Viewers', 'Viewers'])

# Plot pie chart for Viewers by Type
labels = list(type_totals.keys())
sizes = list(type_totals.values())

plt.figure(figsize=(8, 6))
plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
plt.title('Viewers by Popular Type')
plt.axis('equal')  # Equal aspect ratio makes the pie chart circular
plt.tight_layout()

pdf_path = "report_viewers_summary.pdf"

# Function to parse date from filename
def extract_date_from_filename(filename):
    try:
        return datetime.strptime(filename.split('_')[-1].replace('.xlsx', ''), "%d-%b,%Y")
    except Exception:
        return None

# Identify the latest file by date in filename
latest_file = None
latest_date = None
for filename in os.listdir(folder_path):
    if filename.startswith('SiteAnalyticsData_') and filename.endswith('.xlsx'):
        file_date = extract_date_from_filename(filename)
        if file_date and (latest_date is None or file_date > latest_date):
            latest_date = file_date
            latest_file = filename

# Read data from the latest Usage by device sheet
device_usage_data = []
device_heatmaps = {}  # For Page 5

if latest_file:
    file_path = os.path.join(folder_path, latest_file)
    workbook = load_workbook(filename=file_path, data_only=True)
    print(f"\nProcessing 'Usage by device' from latest file: {latest_file}")

    if "Usage by device" in workbook.sheetnames:
        sheet = workbook["Usage by device"]

        # Collect the data and limit it to the last 30 entries
        rows = list(sheet.iter_rows(min_row=2, values_only=True))  # Assuming row 1 is headers
        last_30_rows = rows[-30:]

        # Page 3: Total visits per date
        for row in last_30_rows:
            if row[0] is None:
                continue
            try:
                date = pd.to_datetime(row[0])
                visits = sum(cell if isinstance(cell, (int, float)) else 0 for cell in row[1:])  # sum visits from all devices
                device_usage_data.append([date, visits])
            except Exception:
                continue

        # Page 5: Device-specific heatmaps
        device_breakdown = {
            'Desktop': 1,
            'Mobile Display': [2, 3],
            'Tablet': 4,
            'Other Devices': 5
        }

        device_data = defaultdict(list)

        for row in last_30_rows:
            if row[0] is None:
                continue
            try:
                date = pd.to_datetime(row[0])
            except Exception:
                continue

            for device, col in device_breakdown.items():
                if isinstance(col, list):
                    value = sum(row[i] if isinstance(row[i], (int, float)) else 0 for i in col)
                else:
                    value = row[col] if isinstance(row[col], (int, float)) else 0
                device_data[device].append((date, value))

        # Build heatmap pivot for each device
        for device, entries in device_data.items():
            df_device_temp = pd.DataFrame(entries, columns=["Date", "Visits"])
            df_device_temp.sort_values("Date", inplace=True)
            df_device_temp.set_index("Date", inplace=True)
            df_device_temp['Week'] = df_device_temp.index.to_series().dt.to_period('W').astype(str)
            df_device_temp['Day'] = df_device_temp.index.to_series().dt.day_name()
            pivot = df_device_temp.pivot_table(index='Day', columns='Week', values='Visits', aggfunc='sum')
            pivot = pivot.reindex(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])
            device_heatmaps[device] = pivot

# Create DataFrame and Heatmap if data exists
if device_usage_data:
    df_device = pd.DataFrame(device_usage_data, columns=["Date", "Total Visits"])
    df_device.sort_values("Date", inplace=True)
    df_device.set_index("Date", inplace=True)

    # Create pivot for heatmap: Assuming we want a week-wise heatmap
    df_device['Week'] = df_device.index.to_series().dt.to_period('W').astype(str)
    df_device['Day'] = df_device.index.to_series().dt.day_name()
    pivot = df_device.pivot_table(index='Day', columns='Week', values='Total Visits', aggfunc='sum')
    pivot = pivot.reindex(['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday'])

    # Plot heatmap
    fig3, ax3 = plt.subplots(figsize=(12, 6))
    sns.heatmap(pivot, cmap="YlOrRd", linewidths=.5, annot=True, fmt=".0f", ax=ax3)
    ax3.set_title('Visits Heatmap by Weekday', fontsize=14)

# Now begin saving to PDF
with PdfPages(pdf_path) as pdf:
    # --- Page 1: Pie Chart ---
    fig1, ax1 = plt.subplots(figsize=(8, 6))
    ax1.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
    ax1.set_title('Viewers by Popular Type')
    ax1.axis('equal')
    pdf.savefig(fig1)
    plt.close(fig1)

    # --- Page 2: Table (plotted using matplotlib) ---
    fig2, ax2 = plt.subplots(figsize=(10, len(df) * 0.4 + 1))
    ax2.axis('off')
    ax2.set_title("Aggregated Content View Summary", fontsize=14, fontweight='bold')

    table = ax2.table(
        cellText=df.values,
        colLabels=df.columns,
        loc='center',
        cellLoc='left',
        colColours=['#CCCCCC'] * len(df.columns)
    )

    table.auto_set_font_size(False)
    table.set_fontsize(9)
    table.scale(1.5, 1.5)
    table.auto_set_column_width(col=list(range(len(df.columns))))

    pdf.savefig(fig2)
    plt.close(fig2)

    # --- Page 3: Heatmap with Horizontal Chunking ---
    if device_usage_data:
        from math import ceil

        max_weeks_per_chart = 5  # Same as Page 5 for consistency
        weeks = pivot.columns.tolist()
        num_chunks = ceil(len(weeks) / max_weeks_per_chart)

        for i in range(num_chunks):
            chunk_weeks = weeks[i * max_weeks_per_chart:(i + 1) * max_weeks_per_chart]
            chunk_pivot = pivot[chunk_weeks]

            fig, ax = plt.subplots(figsize=(3.5 * len(chunk_weeks), 6))
            sns.heatmap(chunk_pivot, cmap="YlOrRd", linewidths=.5, annot=True, fmt=".0f", ax=ax)
            ax.set_title(f'Total Visits Heatmap: Weeks ', fontsize=14)
            plt.tight_layout()

            pdf.savefig(fig)
            plt.close(fig)

    # --- Page 4: Visits by Date Table ---
    if device_usage_data:
        fig4, ax4 = plt.subplots(figsize=(10, len(df_device) * 0.4 + 1))
        ax4.axis('off')
        ax4.set_title("Visits by Date", fontsize=14, fontweight='bold')

        df_table = df_device.reset_index()
        colLabels = df_table.columns

        table = ax4.table(
            cellText=df_table.values,
            colLabels=colLabels,
            loc='center',
            cellLoc='left',
            colColours=['#CCCCCC'] * len(colLabels)
        )

        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1.5, 1.5)

        for i, col in enumerate(df_table.columns):
            table.auto_set_column_width([i])

        date_column_index = df_table.columns.get_loc('Date')
        table.auto_set_column_width([date_column_index])

        pdf.savefig(fig4)
        plt.close(fig4)

    # --- Page 5: Device-specific Heatmaps (horizontally split by weeks) ---
    if device_heatmaps:
        from math import ceil

        max_weeks_per_chart = 5  # You can change this based on how wide you want each chart
        color_maps = {
            'Desktop': "Blues",
            'Mobile Display': "Greens",
            'Tablet': "Oranges",
            'Other Devices': "Purples"
        }

        all_figures = []

        for device, pivot in device_heatmaps.items():
            weeks = pivot.columns.tolist()
            num_chunks = ceil(len(weeks) / max_weeks_per_chart)

            for i in range(num_chunks):
                chunk_weeks = weeks[i * max_weeks_per_chart:(i + 1) * max_weeks_per_chart]
                chunk_pivot = pivot[chunk_weeks]

                fig, ax = plt.subplots(figsize=(3.5 * len(chunk_weeks), 6))
                sns.heatmap(chunk_pivot, cmap=color_maps.get(device, "YlOrRd"), linewidths=.5, annot=True, fmt=".0f", ax=ax)
                ax.set_title(f'{device} Visits Heatmap: Weeks', fontsize=14)
                plt.tight_layout()

                all_figures.append(fig)

        for fig in all_figures:
            pdf.savefig(fig)
            plt.close(fig)

    # --- Page 6: Pie Chart for "30 day report" by Timeframe ---
    if "Usage by time" in workbook.sheetnames:
        sheet = workbook["Usage by time"]
        
        # Extract the data from the 3rd column ("30 day report")
        time_data = []
        for row in sheet.iter_rows(min_row=8, values_only=True):  # Assuming row 1 is headers
            if all(cell is None for cell in row):
                break
            hour_label = row[0]  # Hour label
            day_7 = row[1]  # 7-day data
            day_30 = row[2]  # 30-day data
            
            if isinstance(day_30, (int, float)):
                time_data.append([hour_label, day_30])
        
        # Categorize time data into defined time frames
        timeframes = {
            'Early Morning': 0,
            'Morning': 0,
            'Afternoon': 0,
            'Night': 0,
            'Midnight': 0
        }

        for hour_label, value in time_data:
            # Extract hour from the label (assumes the format is 'Day hh AM/PM')
            try:
                hour = int(hour_label.split(' ')[1].split(':')[0])
            except ValueError:
                continue

            # Assign values to the appropriate time frame
            if 0 <= hour < 5:
                timeframes['Early Morning'] += value
            elif 5 <= hour < 12:
                timeframes['Morning'] += value
            elif 12 <= hour < 18:
                timeframes['Afternoon'] += value
            elif 18 <= hour < 21:
                timeframes['Night'] += value
            elif 21 <= hour < 24:
                timeframes['Midnight'] += value

        # Remove timeframes with a value of 0
        timeframes = {key: value for key, value in timeframes.items() if value > 0}

        # Plot pie chart
        labels = list(timeframes.keys())
        sizes = list(timeframes.values())

        fig7, ax7 = plt.subplots(figsize=(8, 6))
        wedges, texts, autotexts = ax7.pie(
            sizes,
            labels=labels,
            autopct='%1.1f%%',
            startangle=140,
            colors=sns.color_palette("Set3", len(labels))
        )

        # Adjust the placement of the legend to be below the chart
        ax7.set_title("Distribution of Views by Timeframe (30 Day Report)", fontsize=14)
        plt.axis('equal')  # Equal aspect ratio ensures that pie chart is circular.

        # Add the legend below the chart
        # Format legend labels to include the number of items (sizes)
        legend_labels = [f"{label} ({size})" for label, size in zip(labels, sizes)]
        ax7.legend(wedges, legend_labels, title="Timeframes", loc="upper center", bbox_to_anchor=(0.5, -0.15), ncol=3)

        # Adjust layout to make room for the legend
        plt.tight_layout()

        # Save to PDF
        pdf.savefig(fig7)
        plt.close(fig7)

    print(f"\nPage 7: Pie chart for '30 day report' by timeframe added to the PDF.")
